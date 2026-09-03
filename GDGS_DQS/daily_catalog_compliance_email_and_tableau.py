# keep your existing ALATION_TOKEN
# %pip install selenium pandas openpyxl pywin32

import os
import re
import time
import logging
from pathlib import Path
from datetime import datetime

import pandas as pd
import win32com.client as win32
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException

# ---------------- CONFIG ----------------
ALATION_BASE_URL = "https://generalmills.alationcloud.com"

QUERY_IDS = [6071, 6125, 6121, 6215]

QUERY_CONFIGS = [
    {
        "query_id": query_id,
        "query_url": f"{ALATION_BASE_URL}/app/query/{query_id}/overview",
    }
    for query_id in QUERY_IDS
]

QUERY_BEHAVIOR = {
    6121: {"statement_no": 10},
    6071: {},
    6125: {},
    6215: {},
}

CHROME_USER_DATA_DIR = r"C:\Temp\selenium-chrome-profile"
CHROME_PROFILE_DIR = "Default"
DOWNLOAD_DIR = r"C:\Temp\alation_downloads"

LOG_DIR = r"C:\Users\G719851\OneDrive - General Mills\Catalog compliance"
LOG_FILE = os.path.join(LOG_DIR, "catalog_compliance_run.log")
RUN_TRACKER_FILE = os.path.join(LOG_DIR, "catalog_compliance_run_tracker.xlsx")

PAGE_LOAD_WAIT_SEC = 30
CLICK_WAIT_SEC = 30
CHECK_INTERVAL_SEC = 30
RECENT_MINUTES_THRESHOLD = 5
NEW_CSV_WAIT_SEC = 180

RUN_QUERY_XPATH = "//button[normalize-space()='Run Query' or .//span[normalize-space()='Run Query'] or contains(., 'Run Query')]"
RESULT_TIME_XPATH = "//p[contains(normalize-space(), 'Result on ')]"
RUN_ANYWAY_XPATH = (
    "//button[normalize-space()='Run Anyway' "
    "or .//span[normalize-space()='Run Anyway'] "
    "or contains(., 'Run Anyway')]"
)

DOWNLOAD_SELECTORS_DEFAULT = [
    (
        By.XPATH,
        "//*[contains(., 'Quick Run Results')]//following::*[normalize-space()='Download'][1]",
    ),
    (By.XPATH, "//button[normalize-space()='Download']"),
    (By.XPATH, "//a[normalize-space()='Download']"),
    (By.XPATH, "//*[normalize-space()='Download']"),
]

DOWNLOAD_SELECTORS_6121 = [
    (
        By.XPATH,
        "//*[normalize-space()='Statement 10']/ancestor::*[self::div or self::section][1]//*[normalize-space()='Download']",
    ),
    (By.XPATH, "(//*[normalize-space()='Download'])[last()]"),
    (
        By.XPATH,
        "//*[contains(., 'Quick Run Results')]//following::*[normalize-space()='Download'][1]",
    ),
    (By.XPATH, "//button[normalize-space()='Download']"),
    (By.XPATH, "//a[normalize-space()='Download']"),
    (By.XPATH, "//*[normalize-space()='Download']"),
]

SAVE_CONFIG = {
    6071: {
        "excel_name": "Project Contacts.xlsx",
        "sheet_name": "Result#126363",
        "save_dir": r"\\genmills.com\corporate\Tableau_Data\All_Company\Global_Data_Governance_Services",
    },
    6125: {
        "excel_name": "Dataset_Completeness.xlsx",
        "sheet_name": "Result#126364",
        "save_dir": r"\\genmills.com\corporate\Tableau_Data\All_Company\Global_Data_Governance_Services",
    },
    6121: {
        "excel_name": "Main_output.xlsx",
        "sheet_name": "Result#126362",
        "save_dir": r"\\genmills.com\corporate\Tableau_Data\All_Company\Global_Data_Governance_Services",
    },
    6215: {
        "excel_name": f"Catalog Compliance Reminder Details - {datetime.now().strftime('%m%d%Y')}.xlsx",
        "sheet_name": "Data",
        "save_dir": r"C:\Users\G719851\OneDrive - General Mills\Catalog compliance",
    },
}

TRACKER_COLUMNS = {
    query_id: (
        f"{os.path.splitext(config['excel_name'])[0]}({query_id})"
        if query_id != 6215
        else "Catalog Compliance Reminder Details(6215)"
    )
    for query_id, config in SAVE_CONFIG.items()
}


GREEN_FILL = PatternFill(fill_type="solid", start_color="92D050", end_color="92D050")
ORANGE_FILL = PatternFill(fill_type="solid", start_color="FFC000", end_color="FFC000")
RED_FILL = PatternFill(fill_type="solid", start_color="FF0000", end_color="FF0000")

# ---------------- LOGGING ----------------
os.makedirs(LOG_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, mode="a", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)


# ---------------- BROWSER SETUP ----------------
def clear_old_downloads():
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    for pattern in ("*.csv", "*.crdownload"):
        for file_path in Path(DOWNLOAD_DIR).glob(pattern):
            try:
                file_path.unlink()
            except Exception as e:
                logger.warning(f"Could not delete old download {file_path}: {e}")


def build_driver():
    clear_old_downloads()

    options = Options()
    options.add_argument(f"--user-data-dir={CHROME_USER_DATA_DIR}")
    options.add_argument(f"--profile-directory={CHROME_PROFILE_DIR}")
    options.add_argument("--start-maximized")

    prefs = {
        "download.default_directory": DOWNLOAD_DIR,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
    }
    options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(PAGE_LOAD_WAIT_SEC)
    return driver


# ---------------- GENERIC HELPERS ----------------
def wait_clickable(driver, xpath, timeout=CLICK_WAIT_SEC):
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((By.XPATH, xpath))
    )


def wait_present(driver, xpath, timeout=CLICK_WAIT_SEC):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.XPATH, xpath))
    )


def js_click(driver, elem):
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
    try:
        elem.click()
    except Exception:
        driver.execute_script("arguments[0].click();", elem)


def visible_elements(driver, by, selector):
    return [e for e in driver.find_elements(by, selector) if e.is_displayed()]


def normalize_col_name(col_name):
    return str(col_name).strip().lower()


def normalized_column_map(df):
    return {normalize_col_name(col): col for col in df.columns}


def get_unique_emails_from_first_row(df, email_col):
    if email_col is None or df.empty:
        return []

    raw_value = df.iloc[0][email_col]
    if pd.isna(raw_value):
        return []

    unique_emails = []
    seen = set()

    for part in str(raw_value).split(";"):
        email = part.strip()
        if email and email.lower() not in seen:
            seen.add(email.lower())
            unique_emails.append(email)

    return unique_emails


def extract_days_old(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    match = re.search(r"-?\d+(\.\d+)?", text.replace(",", ""))
    if match:
        return float(match.group())

    return None


def append_run_summary_tracker(
    run_timestamp, query_row_counts, email_drafted, error_message=""
):
    os.makedirs(LOG_DIR, exist_ok=True)

    tracker_row = {
        "timestamp (when query finished)": run_timestamp,
        "Project Contacts(6071)": "",
        "Dataset_Completeness(6125)": "",
        "Main_output(6121)": "",
        TRACKER_COLUMNS[6215]: "",
        "Email Drafted": "TRUE" if email_drafted else "FALSE",
        "error_message": error_message,
    }

    for query_id, row_count in query_row_counts.items():
        col_name = TRACKER_COLUMNS.get(query_id)
        if col_name:
            tracker_row[col_name] = row_count

    new_row_df = pd.DataFrame([tracker_row])

    if os.path.exists(RUN_TRACKER_FILE):
        try:
            existing_df = pd.read_excel(RUN_TRACKER_FILE)

            for col in new_row_df.columns:
                if col not in existing_df.columns:
                    existing_df[col] = ""

            for col in existing_df.columns:
                if col not in new_row_df.columns:
                    new_row_df[col] = ""

            new_row_df = new_row_df[existing_df.columns]
            updated_df = pd.concat([existing_df, new_row_df], ignore_index=True)

        except Exception:
            updated_df = new_row_df
    else:
        updated_df = new_row_df

    with pd.ExcelWriter(RUN_TRACKER_FILE, engine="openpyxl") as writer:
        updated_df.to_excel(writer, index=False, sheet_name="Run Tracker")


# ---------------- ALATION INTERACTION ----------------
def click_run_query(driver):
    elem = wait_clickable(driver, RUN_QUERY_XPATH)
    js_click(driver, elem)


def click_run_anyway_if_present(driver, wait_sec=8, retries=3):
    for attempt in range(retries):
        try:
            elem = WebDriverWait(driver, wait_sec).until(
                EC.element_to_be_clickable((By.XPATH, RUN_ANYWAY_XPATH))
            )
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", elem
            )

            try:
                elem.click()
            except Exception:
                driver.execute_script("arguments[0].click();", elem)

            logger.info("Clicked 'Run Anyway'")
            return True

        except TimeoutException:
            return False

        except StaleElementReferenceException:
            logger.warning(
                f"'Run Anyway' became stale on attempt {attempt + 1}, retrying..."
            )
            time.sleep(1)

    return False


def click_second_run(driver, main_handle, wait_sec=30):
    end_time = time.time() + wait_sec

    while time.time() < end_time:
        handles = driver.window_handles
        if len(handles) > 1:
            driver.switch_to.window(handles[-1])

        buttons = visible_elements(driver, By.XPATH, RUN_QUERY_XPATH)
        if buttons:
            js_click(driver, buttons[-1])
            click_run_anyway_if_present(driver, wait_sec=5)

            if main_handle in driver.window_handles:
                driver.switch_to.window(main_handle)
            return

        time.sleep(2)

    raise RuntimeError("Could not find second Run Query button")


def get_result_time(driver):
    elem = wait_present(driver, RESULT_TIME_XPATH)
    text = " ".join(elem.text.split())
    ts = text.replace("Result on ", "").strip()
    return datetime.strptime(ts, "%B %d, %Y %I:%M %p")


def click_statement(driver, statement_no):
    label = f"Statement {statement_no}"

    possible_xpaths = [
        f"//*[normalize-space()='{label}']",
        f"//div[normalize-space()='{label}']",
        f"//span[normalize-space()='{label}']",
        f"//li[normalize-space()='{label}']",
        f"//button[normalize-space()='{label}']",
        f"//a[normalize-space()='{label}']",
    ]

    last_error = None

    for _ in range(12):
        for xp in possible_xpaths:
            try:
                elems = driver.find_elements(By.XPATH, xp)
                for elem in elems:
                    if elem.is_displayed():
                        js_click(driver, elem)
                        logger.info(f"Clicked {label}")
                        return
            except Exception as e:
                last_error = e

        driver.execute_script("window.scrollBy(0, 500);")
        try:
            body = driver.find_element(By.TAG_NAME, "body")
            body.send_keys(Keys.PAGE_DOWN)
        except Exception:
            pass

        time.sleep(2)

    raise RuntimeError(f"Could not find {label}. Last error: {last_error}")


def click_download(driver, query_id):
    behavior = QUERY_BEHAVIOR.get(query_id, {})
    if behavior.get("statement_no"):
        click_statement(driver, behavior["statement_no"])

    selectors = (
        DOWNLOAD_SELECTORS_6121 if query_id == 6121 else DOWNLOAD_SELECTORS_DEFAULT
    )

    last_error = None
    for by, selector in selectors:
        try:
            elems = driver.find_elements(by, selector)
            for elem in elems:
                if elem.is_displayed():
                    js_click(driver, elem)
                    logger.info(
                        f"Clicked Download for query {query_id} using {selector}"
                    )
                    return
        except Exception as e:
            last_error = e

    raise RuntimeError(
        f"Download button not found for query {query_id}. Last error: {last_error}"
    )


# ---------------- DOWNLOAD HELPERS ----------------
def get_existing_csv_files():
    return {str(p.resolve()) for p in Path(DOWNLOAD_DIR).glob("*.csv")}


def wait_for_new_csv(existing_files, timeout=NEW_CSV_WAIT_SEC):
    end_time = time.time() + timeout
    download_path = Path(DOWNLOAD_DIR)

    while time.time() < end_time:
        current_csvs = {str(p.resolve()): p for p in download_path.glob("*.csv")}
        new_files = [Path(p) for p in current_csvs if p not in existing_files]
        crdownload_files = list(download_path.glob("*.crdownload"))

        if new_files and not crdownload_files:
            latest_csv = max(new_files, key=lambda p: p.stat().st_mtime)

            size1 = latest_csv.stat().st_size
            time.sleep(2)
            size2 = latest_csv.stat().st_size

            if size1 == size2 and size1 > 0:
                return str(latest_csv)

        time.sleep(2)

    raise TimeoutError("New downloaded CSV was not found in time.")


def close_extra_tabs(driver):
    handles = driver.window_handles[:]
    if not handles:
        return

    main_handle = handles[0]

    for handle in handles[1:]:
        try:
            driver.switch_to.window(handle)
            driver.close()
        except Exception:
            pass

    driver.switch_to.window(main_handle)


# ---------------- CSV / EXCEL HELPERS ----------------
def read_csv_safely(csv_path):
    attempts = [
        {"encoding": "utf-8-sig", "sep": None, "engine": "python"},
        {"encoding": "utf-8", "sep": None, "engine": "python"},
        {"encoding": "latin1", "sep": None, "engine": "python"},
    ]

    last_error = None
    for kwargs in attempts:
        try:
            return pd.read_csv(csv_path, **kwargs)
        except Exception as e:
            last_error = e

    raise RuntimeError(f"Could not read CSV: {csv_path}. Last error: {last_error}")


def apply_days_old_formatting(ws, df):
    days_old_col_idx = None

    for i, col in enumerate(df.columns, start=1):
        if normalize_col_name(col) == "days old":
            days_old_col_idx = i
            break

    if days_old_col_idx is None:
        logger.warning("'Days Old' column not found")
        return

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=days_old_col_idx)
        days_old = extract_days_old(cell.value)

        if days_old is None:
            continue
        if 0 <= days_old <= 21:
            cell.fill = GREEN_FILL
        elif 22 <= days_old <= 30:
            cell.fill = ORANGE_FILL
        elif days_old > 30:
            cell.fill = RED_FILL


def transform_6215_dataframe(df):
    col_map = normalized_column_map(df)

    status_col = col_map.get("status_message")
    last_updated_col = col_map.get("last_updated")
    to_email_col = col_map.get("to_email_list")

    to_email_list = get_unique_emails_from_first_row(df, to_email_col)

    cols_to_drop = [
        col for col in [status_col, last_updated_col, to_email_col] if col is not None
    ]

    if cols_to_drop:
        df = df.drop(columns=cols_to_drop)

    return df, to_email_list


def convert_csv_to_xlsx(query_id, csv_path):
    config = SAVE_CONFIG[query_id]
    os.makedirs(config["save_dir"], exist_ok=True)

    output_path = os.path.join(config["save_dir"], config["excel_name"])
    df = read_csv_safely(csv_path)
    to_email_list = []

    if query_id == 6215:
        df, to_email_list = transform_6215_dataframe(df)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=config["sheet_name"])

        if query_id == 6215:
            ws = writer.book[config["sheet_name"]]
            apply_days_old_formatting(ws, df)

    logger.info(f"Saved XLSX for query {query_id}: {output_path}")
    return output_path, to_email_list, len(df)


# ---------------- QUERY EXECUTION ----------------
def process_query(driver, query_id, query_url):
    logger.info(f"Processing query {query_id}")
    driver.set_page_load_timeout(120)
    driver.get(query_url)

    main_handle = driver.current_window_handle

    click_run_query(driver)
    logger.info(f"Query {query_id}: clicked first Run Query")

    click_run_anyway_if_present(driver, wait_sec=5)

    try:
        click_second_run(driver, main_handle)
        logger.info(f"Query {query_id}: clicked second Run Query")
    except RuntimeError:
        logger.info(f"Query {query_id}: second Run Query not found, continuing")

    click_run_anyway_if_present(driver, wait_sec=5)

    while True:
        displayed_time = get_result_time(driver)
        now = datetime.now()
        diff_minutes = abs((now - displayed_time).total_seconds()) / 60

        logger.info(
            f"Query {query_id} | Displayed: {displayed_time} | Now: {now} | Diff mins: {round(diff_minutes, 2)}"
        )

        if diff_minutes <= RECENT_MINUTES_THRESHOLD:
            existing_files = get_existing_csv_files()
            click_download(driver, query_id)
            logger.info(f"Query {query_id}: download triggered")

            csv_path = wait_for_new_csv(existing_files)
            logger.info(f"Query {query_id}: downloaded CSV found at {csv_path}")

            close_extra_tabs(driver)
            driver.get("about:blank")

            return {
                "query_id": query_id,
                "csv_path": csv_path,
            }

        time.sleep(CHECK_INTERVAL_SEC)


# ---------------- OUTLOOK HELPER ----------------
def draft_catalog_compliance_email(to_emails, attachment_path):
    DISPLAY_EMAIL = True
    SEND_EMAIL = False

    FROM_MAILBOX = "data.governance@genmills.com"
    CC_EMAIL = "dlist39399601@genmills.com;rohit.surve@genmills.com;flory.mascarenhas@genmills.com"
    SUBJECT = f"Catalog Compliance Notice – {datetime.now().strftime('%m/%d/%Y')}"

    BODY_HTML = f"""
    <html>
        <body style="font-family:Aptos, Calibri, Arial, sans-serif; font-size:12pt; color:#000000;">
            <p>Dear All,</p>

            <p>
                This communication is directed to you in recognition of your pivotal role as key figures
                in the governance of data within Big Query. In accordance with a recent correspondence,
                we wish to apprise you of the <strong>restart</strong> of the Catalog Compliance process.
            </p>

            <p>
                Attached is a detailed list of Big Query tables that currently lack complete documentation.
                <u>Please note that this list contains all tables from just created (0 days) to those that have
                crossed the 30<sup>th</sup> day mark</u>.
            </p>

            <p>
                <strong>Please note that the
                <a href="https://tableau.generalmills.com/#/views/AuditReportrevised_16445673428040/AuditBreakdownFullHistoryV1_2revised?:iid=2">
                    GCP Audit Report
                </a>
                is now functional and points to the Cloud data. However, the refresh schedule for the report
                is once/day (anytime between 8 to 10 PM IST/ 9:30 to 11:30 AM CST; depending on the ETL runs).</strong>
            </p>

            <p>
                Please refer to the Column <strong>'Days_Old'</strong> within the attached excel for further details.
            </p>

            <table style="border-collapse:collapse; font-family:Aptos, Calibri, Arial, sans-serif; font-size:12pt;" border="1" cellpadding="6">
                <tr style="background-color:#D9D9D9; font-weight:bold; text-align:center;">
                    <td>Days Old</td>
                    <td>Color</td>
                    <td>Compliance Status</td>
                    <td>Details / Recommended Action</td>
                </tr>
                <tr>
                    <td style="text-align:center;">0-21</td>
                    <td style="background-color:#92D050;"></td>
                    <td><b>Early Stage</b></td>
                    <td>Documentation window open. Begin process as soon as possible.</td>
                </tr>
                <tr>
                    <td style="text-align:center;">22-30</td>
                    <td style="background-color:#FFC000;"></td>
                    <td><b>Action Required Soon</b></td>
                    <td>Within compliance window but approaching the 30-day deadline. Expedite documentation.</td>
                </tr>
                <tr>
                    <td style="text-align:center;">&gt; 30</td>
                    <td style="background-color:#FF0000;"></td>
                    <td><b>Overdue / Critical</b></td>
                    <td>Beyond compliance window. Requires immediate attention and escalation.</td>
                </tr>
            </table>

            <p>
                We rely on those listed under <strong>‘Governance Contact’</strong> below to lead the cataloging effort for
                their <strong>respective</strong> Big Query Project and address any missing table documentation. Please note
                that for tables categorized as 'Overdue / Critical' (beyond 30 days), failure to complete
                documentation will result in the revocation of user access 31 days from their creation date.
            </p>

            <ul>
                <li>
                    If there are updates to the Governance Contact list, kindly send the revised information to
                    <a href="mailto:data.governance@genmills.com">data.governance@genmills.com</a>.
                </li>
                <li>
                    Any inquiries or concerns may be directed to
                    <a href="mailto:DLIST39399601@genmills.com">Catalog Compliance DISTLIST</a>.
                </li>
            </ul>

            <p>
                We greatly appreciate your prompt action in this matter and your ongoing commitment to ensuring
                compliance and collaboration on this critical initiative.
            </p>

            <p>
                Requirements governed by
                <a href="https://genmills.service-now.com/esc?id=policy_view&policy=POL0020732">
                    POL0020732 - Data_Catalog_Standard_20260618
                </a>
                and
                <a href="https://genmills.service-now.com/esc?id=policy_view&policy=POL0020726">
                    POL0020726 - Data_Catalog_Procedure_20260619
                </a>
                as part of the
                <a href="https://genmills.service-now.com/esc?id=policy_view&policy=POL0020409">
                    Information and Data Governance Policy (CP-20).
                </a>
            </p>

            <p>
                Thank you for your attention and cooperation.<br>
                Global Data Governance Services
            </p>
        </body>
    </html>
    """

    outlook = win32.Dispatch("Outlook.Application")
    session = outlook.Session
    mail = outlook.CreateItem(0)

    if to_emails:
        mail.To = "; ".join(to_emails)

    mail.CC = CC_EMAIL
    mail.Subject = SUBJECT
    mail.HTMLBody = BODY_HTML
    mail.Attachments.Add(attachment_path)

    selected_account = None
    for account in session.Accounts:
        try:
            if str(account.SmtpAddress).strip().lower() == FROM_MAILBOX.lower():
                selected_account = account
                break
        except Exception:
            pass

    if selected_account is not None:
        try:
            mail._oleobj_.Invoke(*(64209, 0, 8, 0, selected_account))
        except Exception as e:
            logger.warning(f"Could not set sending account directly: {e}")

    try:
        mail.SentOnBehalfOfName = FROM_MAILBOX
    except Exception as e:
        logger.warning(f"Could not set SentOnBehalfOfName: {e}")

    if SEND_EMAIL:
        mail.Send()
        logger.info("Email sent")
        return True
    elif DISPLAY_EMAIL:
        mail.Display()
        logger.info("Email draft opened")
        return True

    return False


# ---------------- MAIN ----------------
def main():
    driver = build_driver()
    downloaded_files = []

    try:
        for config in QUERY_CONFIGS:
            result = process_query(
                driver=driver,
                query_id=config["query_id"],
                query_url=config["query_url"],
            )
            downloaded_files.append(result)

    finally:
        try:
            driver.quit()
            logger.info("Chrome closed")
        except Exception:
            pass

    catalog_notice_attachment = None
    catalog_notice_to_list = []
    query_row_counts = {}
    run_error_messages = []
    email_drafted = False

    for item in downloaded_files:
        query_id = item["query_id"]
        csv_path = item["csv_path"]

        try:
            xlsx_path, to_emails, row_count = convert_csv_to_xlsx(query_id, csv_path)
            query_row_counts[query_id] = row_count

            if query_id == 6215:
                catalog_notice_attachment = xlsx_path
                catalog_notice_to_list = to_emails

            try:
                os.remove(csv_path)
                logger.info(f"Deleted temp CSV: {csv_path}")
            except Exception as e:
                logger.warning(f"Could not delete temp CSV {csv_path}: {e}")

        except Exception as e:
            logger.error(f"Failed converting query {query_id}: {e}")
            run_error_messages.append(f"Query {query_id}: {str(e)}")

    if catalog_notice_attachment:
        email_drafted = draft_catalog_compliance_email(
            to_emails=catalog_notice_to_list,
            attachment_path=catalog_notice_attachment,
        )
    else:
        logger.warning("6215 attachment not available, email draft skipped")

    run_finished_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    combined_error_message = " | ".join(run_error_messages)

    append_run_summary_tracker(
        run_timestamp=run_finished_timestamp,
        query_row_counts=query_row_counts,
        email_drafted=email_drafted,
        error_message=combined_error_message,
    )


if __name__ == "__main__":
    main()
